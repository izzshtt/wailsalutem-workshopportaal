import io
from datetime import date

from flask import abort, render_template, request, send_file
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, extract

from app import db
from app.dashboard import bp
from app.models import Workshop, Registration
from app.decorators import role_required


MINUTES_PER_HOUR = 60
TOP_SCHOOLS_LIMIT = 10


def _build_workshop_query(date_from, date_to):
    """Return a Workshop query filtered by the given date range strings."""
    query = Workshop.query
    if date_from:
        try:
            query = query.filter(
                Workshop.date >= date.fromisoformat(date_from)
            )
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(
                Workshop.date <= date.fromisoformat(date_to)
            )
        except ValueError:
            pass
    return query


def _get_kpi_base_data(date_from='', date_to=''):
    """Return (total_workshops, total_participants, unique_schools)."""
    base = _build_workshop_query(date_from, date_to)

    total_workshops = (
        db.session.query(func.count(Workshop.id))
        .filter(Workshop.id.in_(base.with_entities(Workshop.id)))
        .scalar() or 0
    )
    total_participants = (
        db.session.query(func.sum(Workshop.participant_count))
        .filter(Workshop.id.in_(base.with_entities(Workshop.id)))
        .scalar() or 0
    )
    unique_schools = (
        db.session.query(func.count(func.distinct(Workshop.school)))
        .filter(Workshop.id.in_(base.with_entities(Workshop.id)))
        .scalar() or 0
    )
    return total_workshops, total_participants, unique_schools


def _get_impact_tables(date_from='', date_to=''):
    """Fetch monthly and top-school aggregates, optionally filtered by date range."""
    year_column = extract('year', Workshop.date).label('jaar')
    month_column = extract('month', Workshop.date).label('maand')

    base_ids = _build_workshop_query(date_from, date_to).with_entities(
        Workshop.id
    )

    per_maand = (
        db.session.query(
            year_column,
            month_column,
            func.count(Workshop.id).label('aantal'),
            func.sum(Workshop.participant_count).label('deelnemers'),
        )
        .filter(Workshop.id.in_(base_ids))
        .group_by(year_column, month_column)
        .order_by(year_column, month_column)
        .all()
    )

    top_scholen = (
        db.session.query(
            Workshop.school,
            func.count(Workshop.id).label('aantal'),
            func.sum(Workshop.participant_count).label('deelnemers'),
        )
        .filter(Workshop.id.in_(base_ids))
        .group_by(Workshop.school)
        .order_by(func.sum(Workshop.participant_count).desc())
        .limit(TOP_SCHOOLS_LIMIT)
        .all()
    )

    return per_maand, top_scholen


def _calculate_total_hours(date_from='', date_to=''):
    """Calculate total hours (converted from minutes) for workshops in date range."""
    total_minutes = (
        db.session.query(func.sum(Workshop.duration))
        .filter(
            Workshop.id.in_(
                _build_workshop_query(date_from, date_to)
                .with_entities(Workshop.id)
            )
        )
        .scalar() or 0
    )
    return round(total_minutes / MINUTES_PER_HOUR, 1)


@bp.route('/')
@login_required
def dashboard():
    """Display the admin dashboard with KPI statistics and recent workshops."""
    today = date.today()
    first_day_of_month = today.replace(day=1)

    total_workshops, total_participants, unique_schools = (
        _get_kpi_base_data()
    )
    workshops_this_month, children_this_month, _ = (
        _get_kpi_base_data(
            date_from=first_day_of_month.isoformat(),
            date_to=today.isoformat()
        )
    )

    active_trainers = (
        db.session.query(func.count(func.distinct(Workshop.trainer_id)))
        .scalar() or 0
    )

    recent_workshops = (
        Workshop.query
        .order_by(Workshop.date.desc())
        .limit(5)
        .all()
    )

    recente_aanmeldingen = (
        Registration.query
        .order_by(Registration.created_at.desc())
        .limit(5)
        .all()
    )

    # Workshops met locatie voor de kaart
    kaart_workshops = Workshop.query.filter(
        Workshop.location.isnot(None)
    ).all()

    kaart_data = [
        {
            'naam': w.name,
            'datum': str(w.date) if w.date else None,
            'locatie': w.location,
            'lat': w.latitude,
            'lng': w.longitude,
        }
        for w in kaart_workshops
    ]

    stats = {
        'totaal_workshops': total_workshops,
        'workshops_deze_maand': workshops_this_month,
        'kinderen_bereikt': total_participants,
        'kinderen_deze_maand': children_this_month,
        'unieke_scholen': unique_schools,
        'actieve_trainers': active_trainers,
    }

    return render_template(
        'dashboard/index.html',
        stats=stats,
        recente_workshops=recent_workshops,
        kaart_data=kaart_data,
        recente_aanmeldingen=recente_aanmeldingen,
    )


@bp.route('/impact')
@login_required
def impact():
    # Deze route toont de impact statistieken met filters voor datum
    date_from = request.args.get('datum_van', '')
    date_to = request.args.get('datum_tot', '')

    total_workshops, total_participants, unique_schools = (
        _get_kpi_base_data(date_from, date_to)
    )
    total_hours = _calculate_total_hours(date_from, date_to)
    per_maand, top_scholen = _get_impact_tables(date_from, date_to)

    return render_template(
        'dashboard/impact.html',
        totaal_workshops=total_workshops,
        totaal_deelnemers=total_participants,
        unieke_locaties=unique_schools,
        totale_uren=total_hours,
        per_maand=per_maand,
        top_scholen=top_scholen,
        datum_van=date_from,
        datum_tot=date_to,
    )


def _style_header_row(ws, num_cols):
    """Apply bold + blue fill to the first row of a worksheet."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')


@bp.route('/export-excel')
@role_required(['admin'])  # Alleen admins mogen exporteren
def export_excel():
    """Exporteer workshop- en impactgegevens naar Excel (alleen voor admins)."""

    date_from = request.args.get('datum_van', '')
    date_to = request.args.get('datum_tot', '')

    total_workshops, total_participants, unique_schools = (
        _get_kpi_base_data(date_from, date_to)
    )
    total_hours = _calculate_total_hours(date_from, date_to)
    per_maand, top_scholen = _get_impact_tables(date_from, date_to)

    wb = Workbook()

    # --- Sheet 1: KPI Overzicht ---
    ws_kpi = wb.active
    ws_kpi.title = 'KPI Overzicht'
    ws_kpi.append(['Indicator', 'Waarde'])
    _style_header_row(ws_kpi, 2)
    ws_kpi.append(['Totaal workshops gegeven', total_workshops])
    ws_kpi.append(['Totaal kinderen bereikt', total_participants])
    ws_kpi.append(['Unieke scholen', unique_schools])
    ws_kpi.append(['Totale lesuren', total_hours])
    if date_from:
        ws_kpi.append(['Periode van', date_from])
    if date_to:
        ws_kpi.append(['Periode tot', date_to])
    ws_kpi.column_dimensions['A'].width = 30
    ws_kpi.column_dimensions['B'].width = 15

    # --- Sheet 2: Workshops per maand ---
    ws_maand = wb.create_sheet('Workshops per maand')
    ws_maand.append(['Jaar', 'Maand', 'Aantal workshops', 'Aantal deelnemers'])
    _style_header_row(ws_maand, 4)
    for rij in per_maand:
        ws_maand.append([
            int(rij.jaar),
            int(rij.maand),
            rij.aantal,
            rij.deelnemers or 0,
        ])
    for col in ['A', 'B', 'C', 'D']:
        ws_maand.column_dimensions[col].width = 20

    # --- Sheet 3: Top scholen ---
    ws_scholen = wb.create_sheet('Top scholen')
    ws_scholen.append(['Rang', 'School', 'Aantal workshops', 'Aantal deelnemers'])
    _style_header_row(ws_scholen, 4)
    for i, school in enumerate(top_scholen, start=1):
        ws_scholen.append([i, school.school, school.aantal, school.deelnemers or 0])
    ws_scholen.column_dimensions['A'].width = 8
    ws_scholen.column_dimensions['B'].width = 35
    ws_scholen.column_dimensions['C'].width = 20
    ws_scholen.column_dimensions['D'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    bestandsnaam = f'wailsalutem-rapportage-{date.today().isoformat()}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=bestandsnaam,
    )


@bp.route('/export-pdf')
@role_required(['admin'])  # Alleen admins mogen exporteren
def export_pdf():
    """Exporteer workshop- en impactgegevens naar PDF (alleen voor admins)."""

    date_from = request.args.get('datum_van', '')
    date_to = request.args.get('datum_tot', '')

    total_workshops, total_participants, unique_schools = (
        _get_kpi_base_data(date_from, date_to)
    )
    total_hours = _calculate_total_hours(date_from, date_to)
    per_maand, top_scholen = _get_impact_tables(date_from, date_to)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    stijl_titel = ParagraphStyle(
        'Titel',
        parent=styles['Title'],
        fontSize=20,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=6,
    )
    stijl_sectie = ParagraphStyle(
        'Sectie',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#1F4E79'),
        spaceBefore=14,
        spaceAfter=6,
    )
    stijl_subtitel = ParagraphStyle(
        'Subtitel',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=16,
    )

    tabel_header_stijl = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#EBF3FB')]),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ])

    periode_tekst = ''
    if date_from and date_to:
        periode_tekst = f'Periode: {date_from} t/m {date_to}'
    elif date_from:
        periode_tekst = f'Periode vanaf: {date_from}'
    elif date_to:
        periode_tekst = f'Periode tot: {date_to}'
    else:
        periode_tekst = 'Alle periodes'

    elementen = [
        Paragraph('WailSalutem — Impact Rapportage', stijl_titel),
        Paragraph(f'Gegenereerd op {date.today().strftime("%d-%m-%Y")} · {periode_tekst}', stijl_subtitel),

        Paragraph('KPI Overzicht', stijl_sectie),
        Table(
            [
                ['Indicator', 'Waarde'],
                ['Totaal workshops gegeven', str(total_workshops)],
                ['Totaal kinderen bereikt', str(total_participants)],
                ['Unieke scholen', str(unique_schools)],
                ['Totale lesuren', str(total_hours)],
            ],
            colWidths=[11 * cm, 5 * cm],
            style=tabel_header_stijl,
        ),
    ]

    if per_maand:
        rijen = [['Jaar', 'Maand', 'Workshops', 'Deelnemers']]
        for rij in per_maand:
            rijen.append([
                str(int(rij.jaar)),
                str(int(rij.maand)),
                str(rij.aantal),
                str(rij.deelnemers or 0),
            ])
        elementen += [
            Paragraph('Workshops per maand', stijl_sectie),
            Table(rijen, colWidths=[4 * cm, 4 * cm, 4 * cm, 5 * cm], style=tabel_header_stijl),
        ]

    if top_scholen:
        rijen = [['#', 'School', 'Workshops', 'Deelnemers']]
        for i, school in enumerate(top_scholen, start=1):
            rijen.append([
                str(i),
                school.school,
                str(school.aantal),
                str(school.deelnemers or 0),
            ])
        elementen += [
            Paragraph('Top scholen op bereik', stijl_sectie),
            Table(rijen, colWidths=[1.5 * cm, 9 * cm, 3.5 * cm, 3 * cm], style=tabel_header_stijl),
        ]

    elementen.append(Spacer(1, 1 * cm))
    elementen.append(Paragraph(
        'Dit rapport bevat geen persoonsgegevens van kinderen (AVG-compliant).',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#999999')),
    ))

    doc.build(elementen)
    output.seek(0)

    bestandsnaam = f'wailsalutem-rapportage-{date.today().isoformat()}.pdf'
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=bestandsnaam,
    )